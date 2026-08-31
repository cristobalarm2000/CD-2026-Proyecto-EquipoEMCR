"""
Cruza los INGRESOS y los TÉRMINOS de causas por su ID (columna "CRR CAUSA" /
"CRR IDCAUSA" / "ID_CAUSA") para saber, causa por causa, cuáles INGRESARON,
cuáles TERMINARON y cuáles TODAVÍA NO TERMINAN (siguen en tramitación).

Sirve para cualquier competencia (Familia, Penal, Civil, ...). Filtra opcional-
mente por texto del tribunal (p. ej. una Corte o un juzgado concreto).

Fuentes admitidas (cualquier mezcla):
  - .xlsx  (una hoja con encabezados en la primera fila)
  - .csv   (';' o ',' , UTF-8)
  - .zip   (los oficiales de numeros.pjud.cl; usa el archivo "por Rol")

Uso:

    python src/cruce_ingresos_terminos.py \
        --ingresos data/raw/familia/Ingresos-2023-CSV.zip data/raw/familia/Ingresos-2024-CSV.zip \
        --terminos data/raw/familia/Terminos-2023-CSV.zip data/raw/familia/Terminos-2024-CSV.zip \
        --filtro-corte 30 \
        --salida reports/cruce_familia_valparaiso

Opciones:
  --filtro-corte TEXTO      Solo causas de esa Corte: código ("30") o nombre
                            ("Valparaiso"). Para toda la V Región usa 30.
  --filtro-tribunal TEXTO   Solo causas cuyo TRIBUNAL contenga ese texto.
  --salida CARPETA          Carpeta de salida (por defecto: reports/cruce).
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
import unicodedata
import zipfile
from datetime import date, datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

csv.field_size_limit(1 << 24)


# ---------------------------------------------------------------------------
# Utilidades de normalización y parsing
# ---------------------------------------------------------------------------
def norm(s: object) -> str:
    """Texto -> MAYÚSCULAS sin acentos, sin U+FFFD, solo alfanumérico."""
    t = str(s if s is not None else "").replace("�", "")
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode()
    return "".join(c for c in t.upper() if c.isalnum())


def limpiar(v: object) -> str:
    return str(v).strip() if v is not None else ""


def parse_fecha(v: object) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y",
                "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_anio(v: object) -> int | None:
    if v in (None, ""):
        return None
    if isinstance(v, (datetime, date)):
        return v.year
    s = str(v).strip()
    if s.isdigit() and len(s) == 4:
        return int(s)
    d = parse_fecha(s)
    return d.year if d else None


# ---------------------------------------------------------------------------
# Detección de columnas
# ---------------------------------------------------------------------------
def mapa_columnas(headers: list) -> dict[str, int | None]:
    H = [norm(h) for h in headers]

    def buscar(*predicados):
        for i, h in enumerate(H):
            if any(p(h) for p in predicados):
                return i
        return None

    return {
        # id de causa: "CRR IDCAUSA" (Penal), "ID_CAUSA", "CRR CAUSA" (Familia)
        "id": buscar(lambda h: "IDCAUSA" in h,
                     lambda h: "ID" in h and "CAUSA" in h,
                     lambda h: h in ("CRRCAUSA", "CRRCAUSAS")),
        "ruc": buscar(lambda h: h == "RUC"),
        "rit": buscar(lambda h: h == "RIT"),
        "cod_corte": buscar(lambda h: "CORTE" in h and ("COD" in h or "CDIGO" in h)),
        "corte": buscar(lambda h: h == "CORTE"),
        "cod_tribunal": buscar(lambda h: "TRIBUNAL" in h and ("COD" in h or "CDIGO" in h)),
        "tribunal": buscar(lambda h: h == "TRIBUNAL"),
        "fecha_ingreso": buscar(lambda h: "FECHA" in h and "INGRESO" in h),
        "anio_ingreso": buscar(lambda h: ("ANO" in h or "ANIO" in h) and "INGRESO" in h),
        "fecha_termino": buscar(lambda h: "FECHA" in h and "TERMINO" in h),
        "anio_termino": buscar(lambda h: ("ANO" in h or "ANIO" in h) and "TERMINO" in h),
        "motivo_termino": buscar(lambda h: "MOTIVO" in h and "TERMINO" in h),
        "duracion": buscar(lambda h: "DURACION" in h),
    }


# ---------------------------------------------------------------------------
# Lectura de fuentes (xlsx / csv / zip)
# ---------------------------------------------------------------------------
def _iter_tablas(path: Path):
    """Genera (headers:list, filas:iterable[list], etiqueta:str)."""
    suf = path.suffix.lower()
    if suf == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            it = ws.iter_rows(values_only=True)
            headers = next(it, None)
            if headers:
                yield list(headers), it, f"{path.name}::{hoja}"
        wb.close()
    elif suf == ".zip":
        zf = zipfile.ZipFile(path)
        miembros = [n for n in zf.namelist()
                    if n.lower().endswith(".csv") and "macosx" not in n.lower()]
        # Preferir el archivo "por Rol" (una fila por causa) y descartar "por Materia".
        if any("rol" in n.lower() for n in miembros):
            miembros = [n for n in miembros if "materia" not in n.lower()]
        for n in miembros:
            fh = io.TextIOWrapper(zf.open(n), encoding="utf-8-sig",
                                  errors="replace", newline="")
            r = csv.reader(fh, delimiter=";")
            headers = next(r, None)
            if headers:
                yield headers, r, f"{path.name}::{Path(n).name}"
    else:  # .csv
        with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
            muestra = fh.read(8192)
            fh.seek(0)
            delim = ";" if muestra.count(";") >= muestra.count(",") else ","
            r = csv.reader(fh, delimiter=delim)
            headers = next(r, None)
            if headers:
                yield headers, list(r), path.name


def _mejor_texto(a: str, b: str) -> str:
    """Elige el texto más completo; penaliza el 'mojibake' (U+FFFD)."""
    if not a:
        return b
    if not b:
        return a
    if "�" in a and "�" not in b:
        return b
    return a


def _merge(dst: dict, src: dict) -> None:
    """Combina dos registros de la misma causa quedándose con lo más informativo."""
    for k in ("ruc", "rit", "cod_corte", "corte", "cod_tribunal", "tribunal",
              "motivo_termino", "duracion"):
        if src.get(k):
            dst[k] = _mejor_texto(str(dst.get(k) or ""), str(src[k]))
    for k in ("fecha_ingreso", "fecha_termino"):  # nos quedamos con la más temprana
        if src.get(k) and (not dst.get(k) or src[k] < dst[k]):
            dst[k] = src[k]
    for k in ("anio_ingreso", "anio_termino"):
        if src.get(k) and not dst.get(k):
            dst[k] = src[k]


def cargar(path: Path, filtro_tribunal: str = "",
           filtro_corte: str = "") -> tuple[str, dict[str, dict]]:
    """Devuelve (tipo, {id_causa: registro}). tipo ∈ {'ingresos','terminos'}."""
    registros: dict[str, dict] = {}
    tipo = "ingresos"
    filtro_t = norm(filtro_tribunal)
    filtro_c = norm(filtro_corte)
    filas_ok = filas_fuera = 0

    for headers, filas, etiqueta in _iter_tablas(path):
        cm = mapa_columnas(headers)
        if cm["id"] is None:
            continue
        if cm["fecha_termino"] is not None or cm["anio_termino"] is not None:
            tipo = "terminos"

        for row in filas:
            if cm["id"] >= len(row):
                continue
            rid = limpiar(row[cm["id"]])
            if not rid:
                continue

            def g(campo: str):
                i = cm[campo]
                return row[i] if (i is not None and i < len(row)) else None

            tribunal = limpiar(g("tribunal"))
            cod_corte = limpiar(g("cod_corte"))
            corte = limpiar(g("corte"))
            if filtro_t and filtro_t not in norm(tribunal):
                filas_fuera += 1
                continue
            if filtro_c and filtro_c not in norm(cod_corte) and filtro_c not in norm(corte):
                filas_fuera += 1
                continue
            filas_ok += 1

            fi = parse_fecha(g("fecha_ingreso"))
            ft = parse_fecha(g("fecha_termino"))
            nuevo = {
                "id": rid,
                "ruc": limpiar(g("ruc")),
                "rit": limpiar(g("rit")),
                "cod_corte": cod_corte,
                "corte": corte,
                "cod_tribunal": limpiar(g("cod_tribunal")),
                "tribunal": tribunal,
                "fecha_ingreso": fi,
                "anio_ingreso": parse_anio(g("anio_ingreso")) or (fi.year if fi else None),
                "fecha_termino": ft,
                "anio_termino": parse_anio(g("anio_termino")) or (ft.year if ft else None),
                "motivo_termino": limpiar(g("motivo_termino")),
                "duracion": limpiar(g("duracion")),
            }
            if rid in registros:
                _merge(registros[rid], nuevo)
            else:
                registros[rid] = nuevo

    print(f"  {path.name:40} tipo={tipo:9} causas={len(registros):>7,} "
          f"(filas usadas={filas_ok:,} / fuera del filtro={filas_fuera:,})")
    return tipo, registros


def cargar_varios(paths: list[Path], filtro_tribunal: str = "",
                  filtro_corte: str = "") -> dict[str, dict]:
    total: dict[str, dict] = {}
    for p in paths:
        _, regs = cargar(p, filtro_tribunal, filtro_corte)
        for rid, rec in regs.items():
            if rid in total:
                _merge(total[rid], rec)
            else:
                total[rid] = dict(rec)
    return total


# ---------------------------------------------------------------------------
# Análisis
# ---------------------------------------------------------------------------
ESTADO_TERM = "INGRESADA Y TERMINADA"
ESTADO_PEND = "INGRESADA - EN TRAMITACION (NO TERMINA)"
ESTADO_BACK = "TERMINADA - INGRESO EN PERIODO ANTERIOR"


def analizar(ingresos: dict[str, dict], terminos: dict[str, dict]) -> list[dict]:
    filas: list[dict] = []
    for rid in sorted(set(ingresos) | set(terminos)):
        i = ingresos.get(rid)
        t = terminos.get(rid)
        base = dict(i or {})
        if t:
            _merge(base, t)

        aparece_ing = i is not None
        terminada = t is not None
        anio_ing = base.get("anio_ingreso")
        anio_ter = t.get("anio_termino") if t else None

        if aparece_ing and terminada:
            estado = ESTADO_TERM
        elif aparece_ing and not terminada:
            estado = ESTADO_PEND
        else:
            estado = ESTADO_BACK

        fi = base.get("fecha_ingreso")
        ft = base.get("fecha_termino")
        dur = base.get("duracion")
        if not dur and fi and ft:
            dur = (ft - fi).days

        filas.append({
            "id_causa": rid,
            "ruc": base.get("ruc", ""),
            "rit": base.get("rit", ""),
            "cod_corte": base.get("cod_corte", ""),
            "corte": base.get("corte", ""),
            "cod_tribunal": base.get("cod_tribunal", ""),
            "tribunal": base.get("tribunal", ""),
            "anio_ingreso": anio_ing or "",
            "fecha_ingreso": fi.isoformat() if fi else "",
            "en_archivo_ingresos": "SI" if aparece_ing else "NO",
            "terminada": "SI" if terminada else "NO",
            "anio_termino": anio_ter or "",
            "fecha_termino": ft.isoformat() if ft else "",
            "motivo_termino": base.get("motivo_termino", ""),
            "duracion_dias": dur if dur not in (None, "") else "",
            "estado": estado,
        })
    return filas


def escribir_csv(filas: list[dict], ruta: Path, columnas: list[str] | None = None) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    cols = columnas or list(filas[0].keys())
    with ruta.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, delimiter=";", extrasaction="ignore")
        w.writeheader()
        w.writerows(filas)


def resumen(filas: list[dict]) -> dict:
    from collections import Counter, defaultdict
    est = Counter(f["estado"] for f in filas)
    # cohorte: año de ingreso x resultado
    cohorte: dict = defaultdict(lambda: Counter())
    for f in filas:
        ai = f["anio_ingreso"] or "s/i"
        if f["estado"] == ESTADO_PEND:
            cohorte[ai]["en_tramitacion"] += 1
        elif f["estado"] == ESTADO_TERM:
            cohorte[ai][f"termino_{f['anio_termino'] or 's/i'}"] += 1
    # término de causas con ingreso anterior
    back = Counter(f["anio_termino"] or "s/i" for f in filas if f["estado"] == ESTADO_BACK)
    # pendientes por tribunal
    pend_trib = Counter(f["tribunal"] for f in filas if f["estado"] == ESTADO_PEND)
    motivos = Counter(f["motivo_termino"] for f in filas
                      if f["terminada"] == "SI" and f["motivo_termino"])
    return {
        "estados": est, "cohorte": cohorte, "backlog_por_anio_termino": back,
        "pendientes_por_tribunal": pend_trib, "motivos_termino": motivos,
    }


# ---------------------------------------------------------------------------
# Dashboard HTML
# ---------------------------------------------------------------------------
def escribir_dashboard(ruta: Path, filas: list[dict], res: dict, meta: dict) -> None:
    import json
    from collections import Counter

    est = res["estados"]
    cohorte_labels = sorted(k for k in res["cohorte"])
    cohorte_term = [sum(v for kk, v in res["cohorte"][a].items() if kk.startswith("termino_"))
                    for a in cohorte_labels]
    cohorte_pend = [res["cohorte"][a].get("en_tramitacion", 0) for a in cohorte_labels]
    pend_trib = res["pendientes_por_tribunal"].most_common(15)
    motivos = res["motivos_termino"].most_common(10)

    payload = {
        "meta": meta,
        "estados": {"labels": list(est.keys()), "data": list(est.values())},
        "cohorte": {"labels": cohorte_labels, "terminadas": cohorte_term,
                    "pendientes": cohorte_pend},
        "pendientes_tribunal": {"labels": [k for k, _ in pend_trib],
                                "data": [v for _, v in pend_trib]},
        "motivos": {"labels": [k for k, _ in motivos], "data": [v for _, v in motivos]},
    }
    html = _PLANTILLA.replace("__DATOS__", json.dumps(payload, ensure_ascii=False))
    ruta.write_text(html, encoding="utf-8")


_PLANTILLA = r"""<!doctype html>
<html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Cruce Ingresos vs Términos por ID</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.3/chart.umd.min.js"></script>
<style>
 :root{color-scheme:light dark}*{box-sizing:border-box}
 body{margin:0;font:15px/1.5 system-ui,"Segoe UI",Roboto,sans-serif;background:#f4f5f7;color:#1c1f23}
 header{background:#1b3a5b;color:#fff;padding:20px 26px}h1{margin:0 0 4px;font-size:19px}
 header p{margin:0;opacity:.85;font-size:13px}
 main{max-width:1150px;margin:0 auto;padding:20px 26px 60px}
 .kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px;margin:20px 0 28px}
 .kpi{background:#fff;border:1px solid #e2e5ea;border-radius:10px;padding:14px 16px}
 .kpi span{font-size:12px;text-transform:uppercase;letter-spacing:.04em;color:#5b6470}
 .kpi b{display:block;font-size:23px;margin-top:4px}
 .card{background:#fff;border:1px solid #e2e5ea;border-radius:10px;padding:14px 16px;margin-bottom:18px}
 .card h2{font-size:15px;margin:0 0 10px;border-left:4px solid #1b3a5b;padding-left:9px}
 .wrap{position:relative;height:340px}
 @media(prefers-color-scheme:dark){body{background:#14171a;color:#e7e9ec}
  .kpi,.card{background:#1e2226;border-color:#2c3238}.kpi span{color:#97a0ac}}
</style></head><body>
<header><h1>Cruce de Ingresos vs Términos (por ID de causa)</h1>
<p id="meta"></p></header>
<main>
 <div class="kpis" id="kpis"></div>
 <div class="card"><h2>Estado de las causas (cruce por ID)</h2><div class="wrap"><canvas id="c_estados"></canvas></div></div>
 <div class="card"><h2>Por año de ingreso: terminadas vs. en tramitación</h2><div class="wrap"><canvas id="c_cohorte"></canvas></div></div>
 <div class="card"><h2>Causas en tramitación (no terminadas) por tribunal — top 15</h2><div class="wrap"><canvas id="c_pend"></canvas></div></div>
 <div class="card"><h2>Motivo de término — top 10</h2><div class="wrap"><canvas id="c_mot"></canvas></div></div>
</main>
<script>
const D = __DATOS__;
const nf = new Intl.NumberFormat("es-CL");
document.getElementById("meta").textContent = D.meta.descripcion;
const kp = [
  ["Total causas (únicas)", D.meta.total],
  ["Ingresadas y terminadas", D.meta.terminadas],
  ["Ingresadas en tramitación", D.meta.pendientes],
  ["Terminadas con ingreso anterior", D.meta.backlog],
];
document.getElementById("kpis").innerHTML = kp.map(([k,v])=>
  `<div class="kpi"><span>${k}</span><b>${nf.format(v)}</b></div>`).join("");
const base = {responsive:true,maintainAspectRatio:false,plugins:{legend:{display:true}}};
new Chart(c_estados,{type:"bar",data:{labels:D.estados.labels,
  datasets:[{label:"causas",data:D.estados.data,backgroundColor:["#2e7d5b","#b5651d","#7a4fa3"]}]},
  options:{...base,indexAxis:"y",plugins:{legend:{display:false}}}});
new Chart(c_cohorte,{type:"bar",data:{labels:D.cohorte.labels,datasets:[
  {label:"Terminadas",data:D.cohorte.terminadas,backgroundColor:"#2e7d5b"},
  {label:"En tramitación",data:D.cohorte.pendientes,backgroundColor:"#c62828"}]},
  options:{...base,scales:{x:{stacked:true},y:{stacked:true}}}});
new Chart(c_pend,{type:"bar",data:{labels:D.pendientes_tribunal.labels,
  datasets:[{label:"en tramitación",data:D.pendientes_tribunal.data,backgroundColor:"#c62828"}]},
  options:{...base,indexAxis:"y",plugins:{legend:{display:false}}}});
new Chart(c_mot,{type:"bar",data:{labels:D.motivos.labels,
  datasets:[{label:"causas",data:D.motivos.data,backgroundColor:"#1b3a5b"}]},
  options:{...base,indexAxis:"y",plugins:{legend:{display:false}}}});
</script></body></html>
"""


# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--ingresos", nargs="+", type=Path, required=True)
    ap.add_argument("--terminos", nargs="+", type=Path, required=True)
    ap.add_argument("--salida", default="reports/cruce")
    ap.add_argument("--filtro-tribunal", default="",
                    help="Solo causas cuyo TRIBUNAL contenga este texto (vacío = todo).")
    ap.add_argument("--filtro-corte", default="",
                    help="Solo causas de esa Corte: código ('30') o nombre ('Valparaiso').")
    args = ap.parse_args(argv)

    ing_paths, ter_paths = args.ingresos, args.terminos

    print("INGRESOS:")
    ingresos = cargar_varios([Path(p) for p in ing_paths],
                             args.filtro_tribunal, args.filtro_corte)
    print("TERMINOS:")
    terminos = cargar_varios([Path(p) for p in ter_paths],
                             args.filtro_tribunal, args.filtro_corte)

    filas = analizar(ingresos, terminos)
    res = resumen(filas)

    out = Path(args.salida)
    escribir_csv(filas, out / "causas.csv")
    escribir_csv([f for f in filas if f["estado"] == ESTADO_PEND], out / "pendientes.csv")
    escribir_csv([f for f in filas if f["estado"] == ESTADO_BACK],
                 out / "terminadas_ingreso_anterior.csv")

    # resumen.csv (largo)
    filas_res = [{"indicador": "ESTADO: " + k, "valor": v} for k, v in res["estados"].items()]
    for a in sorted(res["cohorte"]):
        for kk, vv in sorted(res["cohorte"][a].items()):
            filas_res.append({"indicador": f"INGRESO {a} / {kk}", "valor": vv})
    for a, v in sorted(res["backlog_por_anio_termino"].items()):
        filas_res.append({"indicador": f"INGRESO ANTERIOR / termino {a}", "valor": v})
    escribir_csv(filas_res, out / "resumen.csv", ["indicador", "valor"])
    escribir_csv([{"tribunal": k, "en_tramitacion": v}
                  for k, v in res["pendientes_por_tribunal"].most_common()],
                 out / "pendientes_por_tribunal.csv", ["tribunal", "en_tramitacion"])

    est = res["estados"]
    meta = {
        "descripcion": (f"{len(filas):,} causas únicas cruzadas por ID · "
                        f"ingresos: {', '.join(Path(p).name for p in ing_paths)} · "
                        f"términos: {', '.join(Path(p).name for p in ter_paths)}"),
        "total": len(filas),
        "terminadas": est.get(ESTADO_TERM, 0),
        "pendientes": est.get(ESTADO_PEND, 0),
        "backlog": est.get(ESTADO_BACK, 0),
    }
    escribir_dashboard(out / "dashboard_cruce.html", filas, res, meta)

    # ---- consola ----
    print("\n" + "=" * 70)
    print("RESULTADO DEL CRUCE POR ID DE CAUSA")
    print("=" * 70)
    for k, v in est.items():
        print(f"  {k:<42} {v:>8,}")
    print(f"  {'TOTAL causas únicas':<42} {len(filas):>8,}")
    print("\nPor AÑO DE INGRESO:")
    print(f"  {'año':>6} | {'ingresadas':>10} | {'terminadas':>10} | {'en tramitación':>14}")
    for a in sorted(res["cohorte"]):
        term = sum(v for kk, v in res["cohorte"][a].items() if kk.startswith("termino_"))
        pend = res["cohorte"][a].get("en_tramitacion", 0)
        print(f"  {str(a):>6} | {term + pend:>10,} | {term:>10,} | {pend:>14,}")
    if res["backlog_por_anio_termino"]:
        print("\nCausas TERMINADAS que ingresaron en un período anterior "
              "(no están en los archivos de ingresos):")
        for a, v in sorted(res["backlog_por_anio_termino"].items()):
            print(f"  terminadas en {a}: {v:,}")
    print(f"\nArchivos escritos en:  {out.resolve()}")
    for f in ["causas.csv", "pendientes.csv", "terminadas_ingreso_anterior.csv",
              "pendientes_por_tribunal.csv", "resumen.csv", "dashboard_cruce.html"]:
        print(f"   - {f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
